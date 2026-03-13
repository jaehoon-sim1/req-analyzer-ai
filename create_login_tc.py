# -*- coding: utf-8 -*-
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from copy import copy

SRC = r'C:\Users\wogns1028\Desktop\TC 샘플\TestCase_샘플_V2.xlsx'
DST = r'C:\Users\wogns1028\Desktop\TC 샘플\TestCase_로그인기능.xlsx'

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb['TestCase']

# --- Save header styles from existing data ---
header_font = copy(ws['C9'].font)
header_align = copy(ws['C9'].alignment)
header_fill = copy(ws['C9'].fill)
header_border = copy(ws['C9'].border)

data_font = copy(ws['E12'].font)
data_border = copy(ws['E12'].border)
depth_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
proc_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
section_font = copy(ws['E11'].font)
section_fill = copy(ws['E11'].fill)
section_align = Alignment(vertical='center')
story_align = Alignment(horizontal='center', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Clear existing data rows (11 onward) ---
# Unmerge all merged cells in data area
merges_to_remove = []
for mc in ws.merged_cells.ranges:
    if mc.min_row >= 11:
        merges_to_remove.append(mc)
for mc in merges_to_remove:
    ws.unmerge_cells(str(mc))

for row in ws.iter_rows(min_row=11, max_row=ws.max_row, max_col=35):
    for cell in row:
        cell.value = None

# --- Update title ---
ws['C3'] = '로그인 기능'

# --- Define test case data ---
# Each section: (story_id, section_title, [test_cases])
# Each test_case: (depth1, depth2, depth3, depth4, depth5, test_type, precondition, procedure, expected, test_status)

sections = [
    ('LG-001', '로그인 > 화면 구성', [
        ('로그인', '화면 구성', None, None, None, None, None,
         '1. 로그인 화면에 접근한다.',
         '1. 아이디 입력란, 비밀번호 입력란, 로그인 버튼이 표시된다.',
         'NR'),
        ('로그인', '화면 구성', '아이디 입력란', None, None, None, None,
         '1. 아이디 입력란을 확인한다.',
         '1. 아이디 입력이 가능한 텍스트 필드가 표시된다.',
         'NR'),
        ('로그인', '화면 구성', '비밀번호 입력란', None, None, None, None,
         '1. 비밀번호 입력란을 확인한다.',
         '1. 비밀번호 입력이 가능한 텍스트 필드가 표시되며, 입력 시 마스킹(*) 처리된다.',
         'NR'),
        ('로그인', '화면 구성', '로그인 버튼', None, None, None, None,
         '1. 로그인 버튼을 확인한다.',
         '1. 로그인 버튼이 표시된다.',
         'NR'),
    ]),
    ('LG-002', '로그인 > 정상 로그인', [
        ('로그인', '정상 로그인', None, None, None, None,
         '유효한 아이디와 비밀번호가 존재',
         '1. 유효한 아이디를 입력한다.\n2. 유효한 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 로그인이 성공하여 메인 화면으로 이동한다.',
         'NR'),
    ]),
    ('LG-003', '로그인 > 유효성 검증', [
        ('로그인', '유효성 검증', '아이디 미입력', None, None, None, None,
         '1. 아이디를 입력하지 않는다.\n2. 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         "1. '땡떙땡' 안내문구가 노출된다.",
         'NR'),
        ('로그인', '유효성 검증', '비밀번호 미입력', None, None, None, None,
         '1. 아이디를 입력한다.\n2. 비밀번호를 입력하지 않는다.\n3. 로그인 버튼을 클릭한다.',
         "1. '땡떙땡' 안내문구가 노출된다.",
         'NR'),
        ('로그인', '유효성 검증', '전체 미입력', None, None, None, None,
         '1. 아이디와 비밀번호를 모두 입력하지 않는다.\n2. 로그인 버튼을 클릭한다.',
         "1. '땡떙땡' 안내문구가 노출된다.",
         'NR'),
        ('로그인', '유효성 검증', '잘못된 아이디', None, None, None,
         '유효한 비밀번호가 존재',
         '1. 존재하지 않는 아이디를 입력한다.\n2. 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 로그인 실패 안내문구가 노출된다.',
         'NR'),
        ('로그인', '유효성 검증', '잘못된 비밀번호', None, None, None,
         '유효한 아이디가 존재',
         '1. 유효한 아이디를 입력한다.\n2. 잘못된 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 로그인 실패 안내문구가 노출된다.',
         'NR'),
    ]),
    ('LG-004', '로그인 > 비밀번호 오류 잠금', [
        ('로그인', '비밀번호 오류 잠금', '1회 오류', None, None, None,
         '유효한 아이디가 존재',
         '1. 유효한 아이디를 입력한다.\n2. 잘못된 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 로그인 실패 안내문구가 노출된다.\n2. 계정은 잠기지 않는다.',
         'NR'),
        ('로그인', '비밀번호 오류 잠금', '4회 연속 오류', None, None, None,
         '유효한 아이디가 존재',
         '1. 유효한 아이디를 입력한다.\n2. 잘못된 비밀번호를 4회 연속 입력한다.',
         '1. 4회째 로그인 실패 안내문구가 노출된다.\n2. 계정은 잠기지 않는다.',
         'NR'),
        ('로그인', '비밀번호 오류 잠금', '5회 오류 시 잠금', None, None, None,
         '유효한 아이디가 존재',
         '1. 유효한 아이디를 입력한다.\n2. 잘못된 비밀번호를 5회 연속 입력한다.',
         '1. 5회째 오류 시 계정이 잠금 처리된다.\n2. 잠금 안내문구가 노출된다.',
         'NR'),
        ('로그인', '비밀번호 오류 잠금', '잠금 후 정상 비밀번호 입력', None, None, None,
         '비밀번호 5회 오류로 계정이 잠금 상태',
         '1. 잠금된 계정의 아이디를 입력한다.\n2. 올바른 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 계정 잠금 안내문구가 노출된다.\n2. 로그인이 불가하다.',
         'NR'),
        ('로그인', '비밀번호 오류 잠금', '4회 오류 후 정상 로그인', None, None, None,
         '비밀번호 4회 오류 상태',
         '1. 유효한 아이디를 입력한다.\n2. 올바른 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 로그인이 성공하여 메인 화면으로 이동한다.\n2. 오류 횟수가 초기화된다.',
         'NR'),
        ('로그인', '비밀번호 오류 잠금', '5회 오류 후 재시도 (잘못된 비밀번호)', None, None, None,
         '비밀번호 5회 오류로 계정이 잠금 상태',
         '1. 잠금된 계정의 아이디를 입력한다.\n2. 잘못된 비밀번호를 입력한다.\n3. 로그인 버튼을 클릭한다.',
         '1. 계정 잠금 안내문구가 노출된다.\n2. 로그인이 불가하다.',
         'NR'),
    ]),
]

def apply_cell_style(cell, font=None, alignment=None, border=None, fill=None):
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill

def apply_border_to_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

# --- Write test cases ---
current_row = 11
tc_num = 1

for story_id, section_title, test_cases in sections:
    # Section header row
    ws.cell(row=current_row, column=3, value=story_id)
    apply_cell_style(ws.cell(row=current_row, column=3),
                     font=Font(name='맑은 고딕', size=9), alignment=story_align, border=thin_border)

    ws.cell(row=current_row, column=5, value=section_title)
    apply_cell_style(ws.cell(row=current_row, column=5),
                     font=Font(name='맑은 고딕', size=10, bold=True),
                     alignment=section_align, fill=section_fill)

    # Merge section title E:AF and apply border
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=32)
    apply_border_to_range(ws, current_row, current_row, 3, 32)

    current_row += 1

    # Data rows
    for tc in test_cases:
        depth1, depth2, depth3, depth4, depth5, test_type, precond, procedure, expected, status = tc
        r = current_row

        # TC No.
        tc_no = f'LG-{tc_num:03d}'
        # Don't write TC No in D column per sample pattern (sample doesn't have it in data rows)

        # Depth columns
        if depth1:
            c = ws.cell(row=r, column=5, value=depth1)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)
        if depth2:
            c = ws.cell(row=r, column=6, value=depth2)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)
        if depth3:
            c = ws.cell(row=r, column=7, value=depth3)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)
        if depth4:
            c = ws.cell(row=r, column=8, value=depth4)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)
        if depth5:
            c = ws.cell(row=r, column=9, value=depth5)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)

        # Test Type (J)
        if test_type:
            c = ws.cell(row=r, column=10, value=test_type)
            apply_cell_style(c, font=data_font, alignment=depth_align, border=thin_border)

        # Precondition (K:M merged)
        if precond:
            c = ws.cell(row=r, column=11, value=precond)
            apply_cell_style(c, font=data_font, alignment=proc_align, border=thin_border)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

        # Procedure (N:R merged)
        if procedure:
            c = ws.cell(row=r, column=14, value=procedure)
            apply_cell_style(c, font=data_font, alignment=proc_align, border=thin_border)
        ws.merge_cells(start_row=r, start_column=14, end_row=r, end_column=18)

        # Expected Results (S:U merged)
        if expected:
            c = ws.cell(row=r, column=19, value=expected)
            apply_cell_style(c, font=data_font, alignment=proc_align, border=thin_border)
        ws.merge_cells(start_row=r, start_column=19, end_row=r, end_column=21)

        # Actual Results (W:Y merged)
        ws.merge_cells(start_row=r, start_column=23, end_row=r, end_column=25)

        # Test Status (AA)
        if status:
            c = ws.cell(row=r, column=27, value=status)
            apply_cell_style(c, font=Font(name='Arial', size=10, bold=True),
                           alignment=depth_align, border=thin_border)

        # Apply borders to all columns in the row
        apply_border_to_range(ws, r, r, 3, 32)

        # Set row height for data rows with content
        ws.row_dimensions[r].height = 45

        tc_num += 1
        current_row += 1

# Set row heights for section header rows
for story_id, section_title, _ in sections:
    pass  # Already set during writing

wb.save(DST)
print(f'Saved to: {DST}')
print(f'Total TCs: {tc_num - 1}')
